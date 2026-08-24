#!/usr/bin/env python3
import argparse
import base64
import copy
import json
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_ITEM_START = 13
BASE_ITEM_END = 24
BASE_PAYMENT_ROW = 29
PDF_FONT = "Noto Sans CJK TC"
LATIN_SERIF_FONT = "Liberation Serif"

# Compact copy of the company seal from the supplied Neverland XLSX.
# This is only used by the emergency programmatic fallback. The preferred
# path is still to load public/templates/Neverland請款單.xlsx so the original
# workbook drawings/styles are preserved.
STAMP_PNG_BASE64 = "iVBORw0KGgoAAAANSUhEUgAAAH8AAACABAMAAADQV6vsAAAAMFBMVEX///////7//v7+/v7+/f3+/Pz9/Pz47/DguLrDcXWvQUaoLTGnKi6lJSmjICWhGR5xuR4ZAAAUZElEQVR42o1Za2BM19p+1p6dRDUya4dWSzLZmSR6UyKCViIdhJaDExF674e6RbXVq0Q5R+uuoj6NUop8qihpmlYvCDGCCiLCcapEJjuJey57JXwtMrPX92NfZuL8+fJjsi9rvftdz3rf570sQvgsSAAIbWIkGpypErGrzXYGgFDzp80fkasBEDsDX8vEqEwnDY9GkxB9USEJ4Gp1eDSqCFebKYReqkqiKdg9EqoTjQs7fYEEf9sDKgDOJEkFGJVU3kwBCeAMMH8Beq8YykBSPiQhP1NANXSzm4P0KwlQJUgBylOAMgkqZwAkTWk+JJJ4hXd06ivlzBysEgqgSnH2bmryqAYMKmWAypn1nc4DCETYYSvTNeAC1wfqr4218ADVeVP1/oBFtPtGouKEZrQ/loP//9+4BA9nnAEgXWslJoYzYXeOE4IMSWqocIbvGEcBxom9mUOlfe1Xz3raoKcdf7FdXoV+PZaHQZTYzVtphcDFOMc6DLl/x6ylCeXW6LSgsTe6e9tqIKsvvG9ctiKMCQI6eBnGDo6t3CjPKirMSlxb7recXbO1K3PS7hGAuhku/60oMQpgJzAtKjv2SzIkqpl/auns67zxgyPh/4mDNaCFigCBEPwJZ/1GXBfTju9pGT3xTeOtoOHS3LDGjekFbSc7PvdZhgGRU42B9Is4PfKwjMptefe908RNvIY0hXedFvLhsviKNgJUwRDAOCCCOaBEqaENX0czzbF1Sqe6Hwe49ffBRQCynLL0UaAGinSKp4be2gcAaIYISlfC5enfZ0UXpQvtsm52qKLoI+OHO6qJM85V7bpe4bbmC0G4/kqe6BU03SEEOzgAtfrZ6JObP8PVxbCvMgT8S5089AN6fpl2pG+LXwHbQMq2LVm0toeJgQ4WmPuD02nnhG0jf0BMpQHQxoyGHZkIjqWwMIjt5ftl6x/z0yXCl1cBsDOxhUUCICADhvlCbrVuCFINBeCNqq2ZJ0xiRLgwqtD4/MWLwQu0xIQKwie8BQAcAjfNhqcB9y+e1nzK3AREAtDW/cAfrzU18E1bkvt6jdMdefLM95GGIQHQAMK4UghoZyCpmsUYeGhJuG8mKf4fY9tEMXWwvdxe/MwGhjseSwABQCk/82bFUeEDIp9Oz7cEDI3GH5BOjzRWICC94bteF/c/mQ0IhuOLBjfh7mVHIW6NJfxSvh/yhiL1eq+eh8xbr/ipVtPx2u5eUy62NCjWLrA0N4DE+tdmTlebmkdNWpploVA2HbgvVDEh0O7OQvCzLYP/7HKj2AAOAgAKAGUkGjTstw0r/uj2meVsN4Dgd7QtgmVFY8c5kz1Vlf12mcjrGFA0Rs+KqZkJFCq7Zyc/ZPo7OmdJklPp7zadQdspBoWtch77zGUwm52Jun3379SnJwe7OfX7Lp3O3rYw6N15NDwXSre9aj2JGA6bYjuS5tONu5kKBt4NJ51RKxE28cDW6MePRZh+d7P7unWnquaPeSveohOnyr1P16QeMfdFADgYFUpzcgqAk198NHH4mk8MWyzk5JepU0/M6HjEIVu+2DE/GcLF8BmyBSIDABLcfc4eGddmZw9e/EmKIT3YHv3o+ElLs59JajEpKKjdiIsQOq4IinMaziQAYArCcWb410DniK2jFo4yfZeQPX/12xjcrUPOPPMRrwlbXTJqp1b8eokMQAIVAZoHCqL96ADkqhMjPvrZH8mu1cyrtg3/arbLFBDx8gFX5S9Lr9YdWV6iGCiA67agKQATN7a8cGKw+TkelLM1ceCa2ekWn8TElhZdXTClX0PYtIJ4UMIgEG7SIyApvr8dvLJ9f6opof6lyvSURRMtUhUvOR778vPXT45eEXV20VmA64Ri4AkZlPhIZPyazKCuCgBoqNsW97fpq/1rqok/OAVqj01cqnn/xqUGACInFtPTjix597B/hWWQaSapND59rOcwrG41bpOSQvfWU9aU3W1zj8NPvTCpiUIEbX4t7wrl+jKOfv1bzccHx9LSQl1EkOKOGcPZb13cACBcTSxYi/F5GFTxRl11z7XTptv1JcSe7vm7HtHn+74gfVKan78mKwC4rf7q5LkDOz+X58koU4CY2pELUsPyxoIlhtvJqbH8X4AIbt9ss8klAMA7/tfi/jhWb9/LFACENj6afeyFA2lDvjy3MsjlxhOZjQu8iGWVCkJ+SiGYXKhABKiNAx0B4fdRN/fmem1uoUcFAAjx8rbt7d4gtt414453Pr8F4q9vF6e4B1cWhSyRrh6rpyoEJ7DxxB1RDNm/VUBQu1bPvDQgThaGuAAEtWtYd3emC+NntxaVN2yfiUGz786c7ELIkvUe9RpCxgPtPSuxseyGDNemXCfE++pLxYAY5vq00rvIBoghm7jWxM+/h01fAMFL1lU3FBVd+EAWUjOme1aKOiPduBoGoNPqEYUBEjrVnfzoGTe8yZl3oQKNmS/NyqLhsT09GMy8rYJjnzzcIFW43QDA914ODMPuN5/48e+FAA5FTgWAkO1r1mzrJaNKOP4t+CoNoJQxQ4PY0Ap4vb+uDwzDaS3frW6NRwV8l6dwBvLD3Us92f07PQ/0PDdPzzupDVTUjfAiAIirXe4AARUP3U5yVwBA6zoAwNFVjuPXG3Jy+K0paCgw7D8QNS8Ot0kE1gqaa3jwu0JamQIAQq5DWQtM5tR74XCERdRg1DB8IWRtYPYExDpAH8hMSM6/pFNyhIK4Kc5K1Y6/ZtSaGYoA8DOGAGHKc21UuKgIf4bVdD+65BVDJaBrjKKsoOg97TWTJQXiTyEdG1LntFEBtpIU+9aFk59LsBKk8NdE4JuNnfs6AVAQiLDimHb+zU7dNrcR8HJXheaOKX/2jzAdXfHcM/8GOIsufv1t3f8hcGKhYXvuxJjawPlB23o7qzNO2W60GDCRaBkE4I+3HPnIZWFgKpDu+znnyLzAvDTpvZE79y+PSt62QjFJmYACYE38kRKANdohwMxQggvQOqwmNKBCiivN2D9tKbFtWGA+SZ6vgAGoq+rRSzCWoGcoQNqCRQIE9dEtfgGti/mNoZ1jLmQmm6u84gw7XMOBmo2MGksXzJjDHnhPwA9Fg/2mID/aFQ/cnfha6SDTPoWaSHq6JwDvU4f80Fm0XKNBOFAXY62Bjv7t4KCjffctLPazMjkQVeEiERCqKLcEEAAgv9kdNrQ+5R5rBaaKc7kf3fx8+O4k/3zE/BH/zkGkye6cAg6A2ZkZmWxCr00OgKqX5lr7sCrl2R9TPV8ofiV5UV1t4qoayNpDe3y6MwngVABAEHk+EvixKtxxyxzf41AhQmunKzBjnffOyZz2lycAQJ1lcTll1wUAsdPvpAIIatd6DekmPq5s75rPvn13MDIsxOJt99XfmAhABuCaUz5PAIgGoHtcsTs1Hq1dv76ypsCQkFLa96tM3s2xf1a+kaFoCRUkCmFf655lZaoAxHNvHxxRKAK8bGR3lp1eAMDm3vTw6PG06cHs0XKmMbwc3j8Ot+MB1ipA9yYh6ealrBku1G3EofcWFbgARC57ePvQ7aPssX1tk+b5vdSlHEsOECCCQ9Aw4gntc5J9fy58zjPaoUnaHEC4FrU795WF1SQuoeJwgIG7H1lvZtOMg+p8IPwSuzY5isTNcWkP1ybG/++H82Vo/NfVg7ZFx9A9n9mfaPHvJb50oI0GADDqsZeSylzNtZTi4O0bEi2jAO58E7V39ogV/Hz7m2XzNL+LRe27RwAVyK6/r/Hajs55fFcFeGlpyEwsBSBEds8YsiELIVsjN4/yB5zKSou7KWEQOYBXu7wYGmETBn6xUgDE0fVL9T2jv7728MAMoePAvbkB5W+WtsK6lvQ0D9vil6QokfGscZQ3Ad6dJb31JLB8wd9ZUP6NPt7SANjTOgcSJ4fAwTC/cUExnszwHnZnlbsA7aQHABJmv5XT+x+xpQP3rw6IN0zr9EJJQAksMABa5ggIV5zl1a7Oi36zstqmqMVzdyUmv9vlWFIAaIfnRAb9w+WvikRKSPesQYUQTnf/JpL5Jt21+EvZVvKa55W0tQW5bRoAvLmOUn9qZxsqoz7iZyBmzLhhilg4+U6X29eM15cG/Oz+uP3U78P+CGxARMjSdUV/IjseqxGBS1QBxKS+X42vL0T5I3yVFSvdwstbJ1UtaNuBiFflBwIwEVWEgQER29cFA7OI3DCmqNgyu5i8AU3b2yQdwI/X39p3K6B6l6AxAFdz2SxJCk/U6GFHvOWqldkTv13Y2rb5IKg/3XZ4/BoQEAq4kiYcTMxAfQHrI1gtKzH56b47SpPcbQV4bx//S+vT2Ou7NnxAO7l7d9jZ1DQbn/a2PjVi99vDZ+V67+1/1C0WRxTCoxs0FYnOynRgvkddhrHNc/8qMTYp5dcvRuaujlDuFcDHHf8lK7opGwDsui8AFY99mgUM8e3EuPVMz7Rif1s4fv27KfcsAMLg+3eE5MbGh97JU8AB5JRVznQBgJjhbDt0WetaDM64Z/6QNIQsyddOFO27HQR5emOeCHAoQFzP/HyTshQAcD4/c3l26j7YfAH5xuj8ouAsZ6LjQALlFywQW6isoLLSoizjRfRX2anFIR9nBXzely986Iwd6HOH7QDOtYJKDKLEogJ7ZbYx9IIbADy/rxzzXRD5IOpV/zaEfBL9gIvvaO5TmgnYdHcUm3rYV1pGMRondiBIVgB4V6Z/B6Ap/asqcx9cSZNJ6Jfs2torB4dQ1LsBTiGqlM0sVABhDFTvToQsped0Z3DqpVZNZJUpv/Zp7l2ejeBB/MH9mlkniRKzG4U5ELIE4UPkvcytANCtlfq2eJHIwzsUAJ7FUmh2nMvtBglwJqOBIAR9DKljOuGsrFtfi3F8qTe3pygJ+QBcbkCe2zp5faVzIngbVuZggI3MAoqKLzLSGxPfMt86ohuvOIo9sxDumw1AcSnrBxWHSTU3DPtihIkAXQmgtdenUapHWAJswinTmaV4CXe3zgpPQOz9HXIUoG5Yj0wZAPNHKpFDB/F4xUt5QFAaZVEnk91AH5nno2xF3aO50uDq6Mazk94BoOB1PkMO50ZDkUosIFvvv3lKE6nfmUZJAwVw4gSm0vAe9ucp1A77OkblOTyAb/83r5d7wzWmtIlMKwHAGZyuVwVqlB4/qCQJ6ac6lDVVg6uCPRsA0OXC793dPKrFZCRqaSB4PMiSpCtu5QkKQBSfjwdaC3p8x+7odub0AECJu1u/D1/1bPYaoU2lIgBZAbSp0VLHwfRzxjjAoPGieg/TQkZmA1NQVe3EPqPJefzlCdFnupu1FadGzSQGZSRSXnQqCzIoAO3O3JlLgRXqUFlbB5fH4kD3ofcHUyZXmEu4ZCxB4+pepl5b6e/YPr10KA2TDtxdhyy+NCD9HVWSUmN5XwWFSAAKaLdffawC4mR/2VaSsDfItmzz0SUkLuI+t5+XTj04IpQZEmqvEyZS5gAA8c7Z1PD26/1lm1YOx7NNj86fTH2Hus3zq1C7cUwPd0VgYOnBKBA0wB1Fjl+ZmSsHcMOARKVb4t7qjgmXA0NbTDPO/keKg9bisTurQhY8Q+dBMbpzCNo+h9V/ez7vvsuHAupJW02KOiOgyysCoLIyrevckHdi05UXW9wHTDvv+pK99kw2hnZsS+qEXOqXa+aI1KB1V+fnH5SGhKm+P8Nifjc2QryaWHYN4nvXt69GQJY3IqX6xDvHlEIA0OIZRAZQHH44Nq51k/rgk11+Om3roGPcS0xbxdstTGjYHnjKI/46w/HPyY/oKjh6MYgUALT9eV33L0P7uj3lxEBFqP/njvfX9Uxoun9CAP8gYnjPxnjf4Dklxf7gCkB7vLrdsthePwk3RlUMZ0Yy9fKHMm1YCh7IP7Z+iWcPOnZFZS7Qe5SWM9VqJFXhE7y1TGEEAGILNm1YTcjReRBFe4CAH7aou9ozx+U5C80sjYCuVKCpFBfzY7T/dgvfgwK4uOyl80k+fgvwkgAIhi88vH/llspntMw30ozmvPlq4L6QpVHNb8p6DeXMnrl4lTuyOTRryWLuL8yCdsd3PZaCjbtw5aldrkBDAlq+DM9QK2S51ghsixY4aq/Y+/ehdz62MBD/liQfy/UioiylesjchTIUq+y7vCl8kn1d0R7FrYGAovrKvGRFEHqq3y79zAJRTN49uvrCCODKCk5OZb5xWTY0kBWAN++oXjY+D5CHMwq0zksvACHSuXlAu2ZjY5NLVzcnD2sF7o77KZL++zmsHgKIHKCAL7r4/Ernlt6qRiUGAIMKAMAuTmbMLFK10txR9uVJbgBlMSOVDp2aEaoYbSBod6tynR75JJAWzgCgGEDX4TXe+sI4cMoACHGzJjRtmK3v+ooMkIa+6oPNTDCihJIV4YGCkJmF4ZbhyTFNjKHGKESJ4t2xPlO/8Q47yFXWODWFmt4IpO4DEERefPxJqwdMiQZFeGn7Ij1O+t58MS/BSDp/2vMncHGHKPu3cR8A+BJ+z/CEMctyog4p3Z7uz3RvFJ7ekmEeHWgDlgL8eqRsYmA+v17pg6PFoIuKJzyn0eXPt0p0gVqJwzp60PbuBSB81LYRBeBSTmKH7l+b9JcT1L3i8LEZHQx/1gJSRmGMxHzcsIMAId4RJydtfMI8DHhovg9ayq7+Afm6U+LN0QCgquHhrRKDyEG8qc1mEaWyFSnVsXFmv2AIU7U7ZbcnGKA0tjgKAFg5T3AKgQi0PPVCIMtuy7sntzya6Y8WZVMpWDRnAKGgQkrLGbLchagNxvf1gxuj1+0/LdaPbfV7SVLJIKYChIbbUU52iQyonWIdKgOc6QIka7bEGUDsBAC4ykCr9aumZlDdF7gn4LSYUE4QeFhMOKFqwNE190gqIAGEgnMm8qga46xaUnXJxqm3CkCCCg6uEhqwohjja4QCoatJu3MOBST6P45Vq6MBqOY87mH6LG78D+f616YR+cXoaPDqgLKQwpiqBgYkc0kqr26oAGap0KobK4D/A3nPk0xGDQDqAAAAAElFTkSuQmCC"


def border_style(color="C8C8C8", style="thin"):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def merge(ws, cell_range, value=None, alignment=None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    if value is not None:
        cell.value = value
    if alignment:
        cell.alignment = alignment
    return cell


def add_company_stamp(ws):
    stamp_stream = BytesIO(base64.b64decode(STAMP_PNG_BASE64))
    stamp = XLImage(stamp_stream)
    stamp.width = 101
    stamp.height = 101
    ws.add_image(stamp, "E35")


def force_font(cell, name=PDF_FONT):
    font = copy.copy(cell.font)
    font.name = name
    cell.font = font


def build_fallback_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "範本"
    ws.sheet_view.showGridLines = False
    widths = {"A": 13, "B": 14, "C": 11, "D": 11, "E": 13, "F": 13, "G": 12, "H": 12, "I": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 38):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 8
    for row in (2, 3, 4):
        ws.row_dimensions[row].height = 27
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[6].height = 44
    ws.row_dimensions[34].height = 54
    ws.row_dimensions[36].height = 86

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    section_fill = PatternFill("solid", fgColor="E7E7E7")
    blue_fill = PatternFill("solid", fgColor="D9E7F2")
    thin = border_style()
    outer_side = Side(style="medium", color="222222")

    for row in range(2, 38):
        for col in range(1, 10):
            cell = ws.cell(row, col)
            cell.border = thin
            cell.font = Font(name=PDF_FONT, size=10)
            cell.alignment = left

    merge(ws, "A2:D4", "Neverland", center).font = Font(name=LATIN_SERIF_FONT, size=19, bold=True)
    merge(ws, "E2:I2", "奈文良多有限公司|統一編號 60343390", left)
    merge(ws, "E3:I3", "負責人: 柯怡安  聯絡電話: 0972-211-049", left)
    merge(ws, "E4:I4", "Email:neverland1332@gmail.com", left)
    merge(ws, "A5:I5", "請     款     單", center).font = Font(name=PDF_FONT, size=12, bold=True)

    ws["A6"] = "訂單日期\n訂單交期"
    ws["A6"].font = Font(name=PDF_FONT, size=10, bold=True)
    ws["A6"].alignment = center
    merge(ws, "B6:D6", "", center)
    ws["E6"] = "類型"
    ws["E6"].font = Font(name=PDF_FONT, size=10, bold=True)
    ws["E6"].alignment = center
    merge(ws, "F6:I6", "", center)
    for col in range(1, 10):
        ws.cell(6, col).fill = blue_fill

    merge(ws, "A7:I7", "客戶資料", left).font = Font(name=PDF_FONT, size=10, bold=True)
    for col in range(1, 10):
        ws.cell(7, col).fill = section_fill
    labels = [(8, "客戶名稱", "統一編號"), (9, "聯絡人", "Email"), (10, "公司地址", "電話/手機")]
    for row, left_label, right_label in labels:
        ws.cell(row, 1).value = left_label
        merge(ws, f"B{row}:D{row}", "", left)
        ws.cell(row, 5).value = right_label
        merge(ws, f"F{row}:I{row}", "", left)

    merge(ws, "A11:I11", "請款內容（NTD)", left).font = Font(name=PDF_FONT, size=10, bold=True)
    for col in range(1, 10):
        ws.cell(11, col).fill = section_fill
    headers = [("A12", "貨號"), ("F12", "建議售價"), ("G12", "經銷價"), ("H12", "數量"), ("I12", "總價（未稅）")]
    merge(ws, "B12:E12", "品項名稱", center)
    for ref, value in headers:
        ws[ref] = value
        ws[ref].alignment = center
    for col in range(1, 10):
        ws.cell(12, col).font = Font(name=PDF_FONT, size=10)
    for row in range(BASE_ITEM_START, BASE_ITEM_END + 1):
        merge(ws, f"B{row}:E{row}", "", left)
        for col in (1, 6, 7, 8, 9):
            ws.cell(row, col).alignment = center

    merge(ws, "A28:I28", "付款資訊", left).font = Font(name=PDF_FONT, size=10, bold=True)
    for col in range(1, 10):
        ws.cell(28, col).fill = section_fill
    ws["A29"] = "費用"
    merge(ws, "B29:C29", 0, center)
    ws["D29"] = "營業稅（5%)"
    merge(ws, "E29:F29", 0, center)
    ws["G29"] = "運費"
    merge(ws, "H29:I29", 0, center)
    ws["A30"] = "請款總金額"
    merge(ws, "B30:I30", "NT$0", center).font = Font(name=PDF_FONT, size=11, bold=True)
    ws["A31"] = "支付方式"
    merge(ws, "B31:D31", "匯款", center)
    ws["E31"] = "匯款日期"
    merge(ws, "F31:I31", "", center)
    merge(ws, "A32:I32", "注意事項", left).font = Font(name=PDF_FONT, size=10, bold=True)
    for col in range(1, 10):
        ws.cell(32, col).fill = section_fill
    ws["A33"] = "甲方"
    merge(ws, "B33:D33", "奈文良多有限公司", center)
    ws["E33"] = "乙方"
    merge(ws, "F33:I33", "", center)
    ws["A34"] = "帳戶資訊"
    merge(ws, "B34:D34", "中國信託（822) 文心分行\n帳號：473541331959", center)
    merge(ws, "E34:I34", "1.此內容為商業機密,雙方有權保密,不得向第三方洩漏報價內容。\n2.收款後於月底前寄出發票", left).font = Font(name=PDF_FONT, size=10, color="FF0000")
    merge(ws, "A35:D35", "客戶/經銷", center)
    merge(ws, "E35:G35", "奈文良多有限公司", center)
    merge(ws, "H35:I35", "業務人員", center)
    ws["A36"] = "負責人/公司\n受權簽章"
    ws["A36"].alignment = left
    merge(ws, "B36:D36", "", center)
    merge(ws, "E36:G36", "", center)
    merge(ws, "H36:I36", "TING", center).font = Font(name=PDF_FONT, size=14)
    merge(ws, "A37:I37", "1.請用印公司大小章、發票章\n2.簽名請用中文正楷\n3.簽署後即視為同意上述所有約定", left)

    for col in range(1, 10):
        ws.cell(2, col).border = Border(top=outer_side, left=ws.cell(2, col).border.left, right=ws.cell(2, col).border.right, bottom=ws.cell(2, col).border.bottom)
        ws.cell(37, col).border = Border(bottom=outer_side, left=ws.cell(37, col).border.left, right=ws.cell(37, col).border.right, top=ws.cell(37, col).border.top)
    for row in range(2, 38):
        ws.cell(row, 1).border = Border(left=outer_side, right=ws.cell(row, 1).border.right, top=ws.cell(row, 1).border.top, bottom=ws.cell(row, 1).border.bottom)
        ws.cell(row, 9).border = Border(right=outer_side, left=ws.cell(row, 9).border.left, top=ws.cell(row, 9).border.top, bottom=ws.cell(row, 9).border.bottom)

    add_company_stamp(ws)
    ws.freeze_panes = None
    ws.print_area = "A2:I37"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    return wb


def copy_item_row(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, 10):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
        dst.number_format = src.number_format
    ws.merge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=5)


def insert_extra_item_rows(ws, count: int):
    if count <= 0:
        return
    shifted_merges = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row > BASE_ITEM_END:
            shifted_merges.append((merged.min_row + count, merged.min_col, merged.max_row + count, merged.max_col))
            ws.unmerge_cells(str(merged))
    ws.insert_rows(BASE_ITEM_END + 1, amount=count)
    for row in range(BASE_ITEM_END + 1, BASE_ITEM_END + 1 + count):
        copy_item_row(ws, BASE_ITEM_END, row)
    for min_row, min_col, max_row, max_col in shifted_merges:
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)


def money(value):
    return round(float(value or 0), 2)


def normalize_dynamic_fonts(ws, item_count: int, extra_rows: int) -> None:
    refs = ["B6", "F6", "B8", "F8", "B9", "F9", "B10", "F10"]
    for ref in refs:
        force_font(ws[ref])
    for row in range(BASE_ITEM_START, BASE_ITEM_START + item_count):
        for col in (1, 2, 6, 7, 8, 9):
            force_font(ws.cell(row, col))
    payment_row = BASE_PAYMENT_ROW + extra_rows
    for ref in (ws.cell(payment_row, 2), ws.cell(payment_row, 4), ws.cell(payment_row, 5), ws.cell(payment_row, 8), ws.cell(payment_row + 1, 2), ws.cell(payment_row + 4, 6)):
        force_font(ref)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template")
    parser.add_argument("--data", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    payload = json.loads(Path(args.data).read_text(encoding="utf-8"))
    template = Path(args.template) if args.template else None
    wb = load_workbook(template) if template and template.exists() else build_fallback_template()
    ws = wb["範本"] if "範本" in wb.sheetnames else wb[wb.sheetnames[0]]

    # The supplied XLSX stores the company seal as a drawing/image. If a future
    # template edit accidentally strips drawings, fail loudly instead of
    # silently exporting a legally incomplete billing form.
    if len(getattr(ws, "_images", [])) == 0:
        add_company_stamp(ws)
    if len(getattr(ws, "_images", [])) == 0:
        raise RuntimeError("請款單模板缺少公司印章圖片")

    ws.title = payload["statementNo"][:31]
    items = payload.get("items", [])
    extra_rows = max(0, len(items) - (BASE_ITEM_END - BASE_ITEM_START + 1))
    insert_extra_item_rows(ws, extra_rows)
    last_template_item_row = BASE_ITEM_END + extra_rows
    payment_row = BASE_PAYMENT_ROW + extra_rows
    total_row = payment_row + 1
    party_row = payment_row + 4

    ws["B6"] = f'{payload["issuedAt"]}\n{payload["periodStart"]} ~ {payload["periodEnd"]}'
    source_label = "經銷寄賣" if payload["sourceType"] == "CONSIGNMENT" else "經銷買斷"
    discount = money(payload["settlementRate"]) * 10
    ws["F6"] = f"{source_label}{discount:g}折稅外加"
    ws["B8"] = payload.get("companyName") or payload.get("channelName") or ""
    ws["F8"] = payload.get("taxId") or ""
    ws["B9"] = payload.get("contactName") or ""
    ws["F9"] = payload.get("contactEmail") or ""
    ws["B10"] = payload.get("billingAddress") or ""
    ws["F10"] = payload.get("contactPhone") or ""

    for row in range(BASE_ITEM_START, last_template_item_row + 1):
        for col in (1, 2, 6, 7, 8, 9):
            ws.cell(row, col).value = None

    for index, item in enumerate(items):
        row = BASE_ITEM_START + index
        ws.cell(row, 1).value = item["sku"]
        name = item["productName"]
        if item.get("size") and item["size"] not in name:
            name = f'{name} {item["size"]}'
        ws.cell(row, 2).value = name
        ws.cell(row, 6).value = money(item["listPrice"])
        ws.cell(row, 7).value = money(item["settlementPrice"])
        ws.cell(row, 8).value = int(item["quantity"])
        ws.cell(row, 9).value = f"=G{row}*H{row}"

    ws.cell(payment_row, 2).value = f"=SUM(I{BASE_ITEM_START}:I{last_template_item_row})"
    tax_percent = money(payload["taxRate"]) * 100
    ws.cell(payment_row, 4).value = f"營業稅（{tax_percent:g}%)"
    ws.cell(payment_row, 5).value = f"=B{payment_row}*{money(payload['taxRate'])}"
    ws.cell(payment_row, 8).value = money(payload.get("shippingFee"))
    ws.cell(total_row, 2).value = f"=B{payment_row}+E{payment_row}+H{payment_row}"
    ws.cell(party_row, 6).value = payload.get("companyName") or payload.get("channelName") or ""
    ws.print_area = f"A2:I{37 + extra_rows}"

    normalize_dynamic_fonts(ws, len(items), extra_rows)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


if __name__ == "__main__":
    main()
